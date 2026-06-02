#!/usr/bin/env python3
"""
Скрипт для метрологической обработки данных из Книга2.xlsx
Река Прохладная — пробы 1-4
Добавляет 2 параллельных измерения к каждой пробе (итого 3 параллельных)
Вычисляет: среднее, S_отн (%), Δ (%)
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import math
import random
import shutil

random.seed(42)  # воспроизводимость

# ─────────────────────────────────────────────────────────────────────────────
# Исходные данные из файла
# ─────────────────────────────────────────────────────────────────────────────

# Структура: проба -> { ион: [конц_п1, конц_п2, конц_п3, конц_п4] }
# Проба 1 (галя а1) — только 4 иона (без нитрита)
# Пробы 2-3 — 5 ионов
# Проба 4 — 4 иона (без нитрита)

# Оригинальные данные (одно измерение)
original_data = {
    "Проба 1": {
        "хлорид":  {"конц": 25.81,   "высота": 14.156, "площадь": 206.8,  "время": 3.143},
        "сульфат": {"конц": 30.89,   "высота": 23.719, "площадь": 201.7,  "время": 3.337},
        "нитрат":  {"конц": 1.382,   "высота": 0.755,  "площадь": 6.894,  "время": 3.537},
        "фосфат":  {"конц": 0.423,   "высота": 0.744,  "площадь": 4.189,  "время": 4.012},
    },
    "Проба 2": {
        "хлорид":  {"конц": 19.59,   "высота": 11.028, "площадь": 156.9,  "время": 3.257},
        "нитрит":  {"конц": 0.03776, "высота": 0.033,  "площадь": 0.236,  "время": 3.403},
        "сульфат": {"конц": 31.38,   "высота": 23.099, "площадь": 204.9,  "время": 3.472},
        "нитрат":  {"конц": 0.8426,  "высота": 0.454,  "площадь": 4.204,  "время": 3.690},
        "фосфат":  {"конц": 0.4257,  "высота": 0.650,  "площадь": 4.216,  "время": 4.237},
    },
    "Проба 3": {
        "хлорид":  {"конц": 21.4,    "высота": 11.483, "площадь": 171.4,  "время": 3.282},
        "нитрит":  {"конц": 0.02384, "высота": 0.020,  "площадь": 0.149,  "время": 3.433},
        "сульфат": {"конц": 33.1,    "высота": 24.324, "площадь": 216.1,  "время": 3.500},
        "нитрат":  {"конц": 1.071,   "высота": 0.528,  "площадь": 5.342,  "время": 3.720},
        "фосфат":  {"конц": 0.4372,  "высота": 0.609,  "площадь": 4.329,  "время": 4.285},
    },
    "Проба 4": {
        "хлорид":  {"конц": 17.46,   "высота": 10.107, "площадь": 139.8,  "время": 3.303},
        "сульфат": {"конц": 27.64,   "высота": 20.859, "площадь": 180.5,  "время": 3.525},
        "нитрат":  {"конц": 1.221,   "высота": 0.582,  "площадь": 6.090,  "время": 3.748},
        "фосфат":  {"конц": 0.422,   "высота": 0.545,  "площадь": 4.178,  "время": 4.332},
    },
}

def gen_parallel(value, cv_frac=0.03):
    """Генерирует параллельное значение с ~3% случайным разбросом."""
    noise = random.gauss(0, cv_frac)
    return round(value * (1 + noise), max(2, len(str(value).rstrip('0').split('.')[-1])))

def round_sig(x, sig=4):
    if x == 0:
        return 0
    d = math.ceil(math.log10(abs(x)))
    power = sig - d
    factor = 10 ** power
    return round(x * factor) / factor

def mean3(a, b, c):
    return (a + b + c) / 3

def s_rel(a, b, c):
    m = mean3(a, b, c)
    if m == 0:
        return 0
    variance = ((a-m)**2 + (b-m)**2 + (c-m)**2) / 2  # n-1=2
    s = math.sqrt(variance)
    return s / m * 100

def delta_percent(a, b, c, t=4.303):
    """
    Δ% = t * S / sqrt(n) / mean * 100
    t = 4.303 при n=3, P=0.95
    """
    m = mean3(a, b, c)
    if m == 0:
        return 0
    variance = ((a-m)**2 + (b-m)**2 + (c-m)**2) / 2
    s = math.sqrt(variance)
    return t * s / math.sqrt(3) / m * 100

# ─────────────────────────────────────────────────────────────────────────────
# Генерация параллельных измерений
# ─────────────────────────────────────────────────────────────────────────────

parallels = {}
for probe, ions in original_data.items():
    parallels[probe] = {}
    for ion, d in ions.items():
        c0 = d["конц"]
        c2 = gen_parallel(c0)
        c3 = gen_parallel(c0)
        parallels[probe][ion] = [c0, c2, c3]

# ─────────────────────────────────────────────────────────────────────────────
# Стили
# ─────────────────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # тёмно-синий
HDR2_FILL  = PatternFill("solid", fgColor="2E75B6")   # синий
PROBE_FILL = PatternFill("solid", fgColor="BDD7EE")   # светло-голубой
RES_FILL   = PatternFill("solid", fgColor="E2EFDA")   # светло-зелёный
ALT_FILL   = PatternFill("solid", fgColor="DEEAF1")   # очень светло-голубой
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HDR2_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
PROBE_FONT = Font(bold=True, color="1F4E79", name="Calibri", size=11)
BODY_FONT  = Font(name="Calibri", size=10)
RES_FONT   = Font(bold=True, name="Calibri", size=10, color="375623")
BOLD_FONT  = Font(bold=True, name="Calibri", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

thin = Side(style="thin", color="8EA9C1")
med  = Side(style="medium", color="1F4E79")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BORDER  = Border(left=med,  right=med,  top=med,  bottom=med)

def apply_border(cell, border=THIN_BORDER):
    cell.border = border

def style_header(cell, fill=HDR_FILL, font=HDR_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    apply_border(cell)

def style_body(cell, fill=WHITE_FILL, font=BODY_FONT, align=CENTER):
    cell.fill = fill
    cell.font = font
    cell.alignment = align
    apply_border(cell)

def style_result(cell):
    cell.fill = RES_FILL
    cell.font = RES_FONT
    cell.alignment = CENTER
    apply_border(cell)

# ─────────────────────────────────────────────────────────────────────────────
# Создание файла
# ─────────────────────────────────────────────────────────────────────────────

shutil.copy(
    "/home/delta/code/course-work/excel/Книга2.xlsx",
    "/home/delta/code/course-work/excel/Книга2_обработанная.xlsx"
)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# Лист 1: Исходные данные (3 параллельных измерения)
# ════════════════════════════════════════════════════════════════════════════

ws_raw = wb.create_sheet("Исходные данные")
ws_raw.sheet_view.showGridLines = False

# Заголовок листа
ws_raw.merge_cells("A1:K1")
c = ws_raw["A1"]
c.value = "Результаты капиллярного электрофореза проб воды р. Прохладная"
c.fill = HDR_FILL
c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
c.alignment = CENTER
ws_raw.row_dimensions[1].height = 30

# Подзаголовок с описанием параллельных
ws_raw.merge_cells("A2:K2")
c = ws_raw["A2"]
c.value = "Три параллельных измерения (П1 — оригинал, П2, П3 — синтезированные параллели при СКО ≈ 3%)"
c.fill = PatternFill("solid", fgColor="BDD7EE")
c.font = Font(italic=True, color="1F4E79", name="Calibri", size=10)
c.alignment = CENTER
ws_raw.row_dimensions[2].height = 18

ROW_START = 4  # строка, с которой начинаются данные пробы

# Ширины столбцов: A(№ иона), B(ион), C-E(конц 1-3), F(среднее), G(S_отн%), H(Δ%)
col_widths = {
    "A": 6,   # №
    "B": 12,  # Ион
    "C": 12,  # П1 конц
    "D": 12,  # П2 конц
    "E": 12,  # П3 конц
    "F": 14,  # Среднее
    "G": 12,  # S_отн %
    "H": 12,  # Δ %
}
for col, w in col_widths.items():
    ws_raw.column_dimensions[col].width = w

current_row = ROW_START

probe_labels = {
    "Проба 1": "Проба 1 (галя а1)",
    "Проба 2": "Проба 2 (галя а5)",
    "Проба 3": "Проба 3 (галя а5)",
    "Проба 4": "Проба 4 (галя а5)",
}

ion_order = {
    "Проба 1": ["хлорид", "сульфат", "нитрат", "фосфат"],
    "Проба 2": ["хлорид", "нитрит", "сульфат", "нитрат", "фосфат"],
    "Проба 3": ["хлорид", "нитрит", "сульфат", "нитрат", "фосфат"],
    "Проба 4": ["хлорид", "сульфат", "нитрат", "фосфат"],
}

for probe_key in ["Проба 1", "Проба 2", "Проба 3", "Проба 4"]:
    label = probe_labels[probe_key]
    ions = ion_order[probe_key]

    # Заголовок пробы
    ws_raw.merge_cells(f"A{current_row}:H{current_row}")
    c = ws_raw[f"A{current_row}"]
    c.value = label
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.alignment = LEFT
    ws_raw.row_dimensions[current_row].height = 20
    apply_border(c, MED_BORDER)
    current_row += 1

    # Заголовки столбцов
    headers = ["№", "Ион", "C, мг/л (П1)", "C, мг/л (П2)", "C, мг/л (П3)",
               "C̄, мг/л", "S_отн, %", "Δ, %"]
    cols = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col, h in zip(cols, headers):
        c = ws_raw[f"{col}{current_row}"]
        c.value = h
        style_header(c, fill=HDR2_FILL, font=HDR2_FONT)
    ws_raw.row_dimensions[current_row].height = 35
    current_row += 1

    # Данные
    for idx, ion in enumerate(ions, start=1):
        d = parallels[probe_key][ion]
        c0, c2, c3 = d[0], d[1], d[2]
        mean_c = round_sig(mean3(c0, c2, c3), 4)
        sr     = round(s_rel(c0, c2, c3), 2)
        dlt    = round(delta_percent(c0, c2, c3), 2)

        fill = WHITE_FILL if idx % 2 == 1 else ALT_FILL

        row_data = [idx, ion, c0, c2, c3, mean_c, sr, dlt]
        for col, val in zip(cols, row_data):
            c = ws_raw[f"{col}{current_row}"]
            c.value = val
            if col in ["F", "G", "H"]:
                style_result(c)
            else:
                style_body(c, fill=fill)
            if col == "B":
                c.alignment = LEFT
        ws_raw.row_dimensions[current_row].height = 18
        current_row += 1

    # Пустая строка
    current_row += 1

# Фиксация первых строк
ws_raw.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# Лист 2: Метрологическая обработка (сводная таблица)
# ════════════════════════════════════════════════════════════════════════════

ws_met = wb.create_sheet("Метрологическая обработка")
ws_met.sheet_view.showGridLines = False

# Ширины
col_widths_met = {
    "A": 14,  # Проба
    "B": 12,  # Ион
    "C": 13,  # C, мг/л (П1)
    "D": 13,  # C, мг/л (П2)
    "E": 13,  # C, мг/л (П3)
    "F": 14,  # C̄
    "G": 12,  # S_отн %
    "H": 12,  # Δ %
}
for col, w in col_widths_met.items():
    ws_met.column_dimensions[col].width = w

# Заголовок
ws_met.merge_cells("A1:H1")
c = ws_met["A1"]
c.value = "Метрологическая обработка результатов анализа проб воды р. Прохладная"
c.fill = HDR_FILL
c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
c.alignment = CENTER
ws_met.row_dimensions[1].height = 30

ws_met.merge_cells("A2:H2")
c = ws_met["A2"]
c.value = ("n = 3 параллельных измерения | t (P=0.95, n=3) = 4.303 | "
           "S_отн = σ/x̄ · 100%  |  Δ = t · S / √n / x̄ · 100%")
c.fill = PatternFill("solid", fgColor="BDD7EE")
c.font = Font(italic=True, color="1F4E79", name="Calibri", size=10)
c.alignment = CENTER
ws_met.row_dimensions[2].height = 16

# Заголовки
headers_met = [
    "Проба", "Ион",
    "C, мг/л\n(П1)", "C, мг/л\n(П2)", "C, мг/л\n(П3)",
    "C̄, мг/л", "S_отн, %", "Δ, %"
]
cols_met = ["A", "B", "C", "D", "E", "F", "G", "H"]
row = 3
for col, h in zip(cols_met, headers_met):
    c = ws_met[f"{col}{row}"]
    c.value = h
    style_header(c)
ws_met.row_dimensions[row].height = 40
row += 1

for probe_key in ["Проба 1", "Проба 2", "Проба 3", "Проба 4"]:
    ions = ion_order[probe_key]
    first = True
    for idx, ion in enumerate(ions):
        d = parallels[probe_key][ion]
        c0, c2, c3 = d[0], d[1], d[2]
        mean_c = round_sig(mean3(c0, c2, c3), 4)
        sr     = round(s_rel(c0, c2, c3), 2)
        dlt    = round(delta_percent(c0, c2, c3), 2)

        fill = WHITE_FILL if idx % 2 == 0 else ALT_FILL

        # Проба (объединить ячейки по кол-ву ионов в пробе)
        if first:
            # Запишем название пробы и объединим ниже после цикла — сделаем иначе:
            # просто запишем в первой строке, остальные оставим пустыми
            c = ws_met[f"A{row}"]
            c.value = probe_key
            c.fill = PROBE_FILL
            c.font = PROBE_FONT
            c.alignment = CENTER
            apply_border(c)
            first = False
        else:
            c = ws_met[f"A{row}"]
            c.value = None
            c.fill = PROBE_FILL
            apply_border(c)

        row_data_met = [None, ion, c0, c2, c3, mean_c, sr, dlt]
        for col, val in zip(cols_met, row_data_met):
            if col == "A":
                continue
            c = ws_met[f"{col}{row}"]
            c.value = val
            if col in ["F", "G", "H"]:
                style_result(c)
            else:
                style_body(c, fill=fill)
            if col == "B":
                c.alignment = LEFT
        ws_met.row_dimensions[row].height = 18
        row += 1

    # Объединяем ячейки столбца A для пробы
    start_probe_row = row - len(ions)
    end_probe_row   = row - 1
    if end_probe_row > start_probe_row:
        ws_met.merge_cells(f"A{start_probe_row}:A{end_probe_row}")
        c = ws_met[f"A{start_probe_row}"]
        c.fill = PROBE_FILL
        c.font = PROBE_FONT
        c.alignment = CENTER
        apply_border(c, MED_BORDER)

    row += 1  # пустая строка между пробами (убираем если не нужна)

ws_met.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# Сохранение
# ════════════════════════════════════════════════════════════════════════════

out_path = "/home/delta/code/course-work/excel/Книга2_обработанная.xlsx"
wb.save(out_path)
print(f"Сохранено: {out_path}")

# Печать контрольной сводки в консоль
print()
print("=" * 90)
print(f"{'Проба':<10} {'Ион':<10} {'П1':>10} {'П2':>10} {'П3':>10} {'Среднее':>12} {'S_отн,%':>10} {'Δ,%':>8}")
print("=" * 90)
for probe_key in ["Проба 1", "Проба 2", "Проба 3", "Проба 4"]:
    for ion in ion_order[probe_key]:
        d = parallels[probe_key][ion]
        c0, c2, c3 = d
        mean_c = round_sig(mean3(c0, c2, c3), 4)
        sr     = round(s_rel(c0, c2, c3), 2)
        dlt    = round(delta_percent(c0, c2, c3), 2)
        print(f"{probe_key:<10} {ion:<10} {c0:>10.5g} {c2:>10.5g} {c3:>10.5g} {mean_c:>12.5g} {sr:>10.2f} {dlt:>8.2f}")
    print("-" * 90)
